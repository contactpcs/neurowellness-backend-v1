#!/usr/bin/env python3
"""
NeuroWellness — PRS Data Excel Generator
Reads all scale JSON files + conditionMap.json and produces a formatted
Excel workbook ready for review or direct DB injection via the seed script.

Usage:
    pip install openpyxl
    python scripts/generate_excel.py

Output:
    neurowellness_prs_data.xlsx  (in project root)
"""

import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, GradientFill
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not found. Run:  pip install openpyxl")
    sys.exit(1)


# ── Paths ──────────────────────────────────────────────────────────────────────
SCRIPT_DIR   = Path(__file__).parent
PROJECT_ROOT = SCRIPT_DIR.parent.parent          # neurowellness/
DATA_ROOT    = PROJECT_ROOT.parent / "existing data" / "data"
SCALES_DIR   = DATA_ROOT / "scales"
CONDITIONS_FILE = DATA_ROOT / "conditionMap.json"
OUTPUT_FILE  = PROJECT_ROOT / "neurowellness_prs_data.xlsx"


# ── Colour palette ─────────────────────────────────────────────────────────────
C = {
    "header_fill":   "4F46E5",   # indigo-600
    "header_font":   "FFFFFF",
    "sub_fill":      "EEF2FF",   # indigo-50
    "alt_fill":      "F9FAFB",   # gray-50
    "border":        "E5E7EB",   # gray-200
    "green_fill":    "DCFCE7",
    "blue_fill":     "DBEAFE",
    "purple_fill":   "F3E8FF",
    "yellow_fill":   "FEF9C3",
    "orange_fill":   "FFEDD5",
    "red_fill":      "FEE2E2",
}


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color="000000", size=10) -> Font:
    return Font(name="Calibri", bold=bold, color=color, size=size)


def _border() -> Border:
    s = Side(style="thin", color=C["border"])
    return Border(left=s, right=s, top=s, bottom=s)


def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def _write_header(ws, headers: list[tuple[str, int]], row=1):
    """Write a styled header row. headers = [(label, col_width), ...]"""
    for col_idx, (label, width) in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=label)
        cell.fill    = _fill(C["header_fill"])
        cell.font    = _font(bold=True, color=C["header_font"], size=10)
        cell.border  = _border()
        cell.alignment = _center()
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[row].height = 22
    ws.freeze_panes = f"A{row + 1}"


def _write_row(ws, row_idx: int, values: list, fill_hex: str | None = None):
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.border = _border()
        cell.alignment = _left()
        cell.font = _font()
        if fill_hex:
            cell.fill = _fill(fill_hex)


# ── ID helpers (mirrors seed_scales.py) ───────────────────────────────────────
def disease_id(disease_name: str) -> str:
    return re.sub(r"[^A-Z0-9/]", "", disease_name.upper().replace("'", "").replace(" ", "")) + "/2026"


def disease_code(disease_name: str) -> str:
    return re.sub(r"[^A-Z0-9]", "", disease_name.upper().replace("'", ""))


def scale_id(code: str) -> str:
    return f"{code}/2026"


def question_id(scale_code: str, idx: int) -> str:
    return f"{scale_code}/{idx + 1:03d}"


def option_id(q_id: str, opt_idx: int) -> str:
    return f"{q_id}/{opt_idx + 1:02d}"


def ds_map_id(disease_name: str, scale_code: str) -> str:
    return f"{disease_name}/{scale_code}"


# ── Load data ──────────────────────────────────────────────────────────────────
def load_all_scales() -> dict[str, dict]:
    """Return { scale_code: scale_dict } for every JSON in SCALES_DIR."""
    scales = {}
    for f in sorted(SCALES_DIR.glob("*.json")):
        with open(f, encoding="utf-8") as fh:
            data = json.load(fh)
        code = data.get("id") or f.stem
        scales[code] = data
    return scales


def load_conditions() -> dict:
    with open(CONDITIONS_FILE, encoding="utf-8") as fh:
        return json.load(fh)


# ── Sheet builders ─────────────────────────────────────────────────────────────

def build_diseases_sheet(wb, conditions: dict):
    ws = wb.create_sheet("Diseases")
    headers = [
        ("disease_id", 38),
        ("disease_code", 22),
        ("disease_name", 32),
        ("description", 55),
        ("version", 10),
        ("status", 10),
        ("scales_count", 14),
    ]
    _write_header(ws, headers)

    fills = [C["green_fill"], C["blue_fill"], C["purple_fill"],
             C["yellow_fill"], C["orange_fill"], C["red_fill"]]

    for i, (cond_id, cond) in enumerate(conditions["conditions"].items()):
        name   = cond["label"]
        d_id   = disease_id(name)
        d_code = disease_code(name)
        fill   = fills[i % len(fills)]
        _write_row(ws, i + 2, [
            d_id,
            d_code,
            name,
            cond.get("description", ""),
            "v1.0",
            True,
            len(cond.get("scales", [])),
        ], fill_hex=fill)

    print(f"  Diseases sheet: {len(conditions['conditions'])} rows")


def build_scales_sheet(wb, scales: dict, conditions: dict):
    ws = wb.create_sheet("Scales")
    headers = [
        ("scale_id", 30),
        ("scale_code", 20),
        ("scale_name", 52),
        ("description", 60),
        ("scoring_type", 16),
        ("max_score", 12),
        ("total_questions", 16),
        ("is_common_scale", 16),
        ("num_diseases_used", 18),
        ("recall_period", 28),
        ("version", 10),
    ]
    _write_header(ws, headers)

    # Count how many diseases use each scale
    scale_usage: dict[str, int] = {}
    for cond in conditions["conditions"].values():
        for sc in cond.get("scales", []):
            scale_usage[sc] = scale_usage.get(sc, 0) + 1

    for i, (code, sc) in enumerate(scales.items()):
        fill = C["sub_fill"] if i % 2 == 0 else None
        total_q = len(sc.get("questions", []))
        _write_row(ws, i + 2, [
            scale_id(code),
            code,
            sc.get("name", code),
            sc.get("description", ""),
            sc.get("scoringType", "sum"),
            sc.get("maxScore", ""),
            total_q,
            False,
            scale_usage.get(code, 0),
            sc.get("recallPeriod", ""),
            sc.get("version", "1.0"),
        ], fill_hex=fill)

    print(f"  Scales sheet: {len(scales)} rows")


def build_disease_scale_map_sheet(wb, conditions: dict):
    ws = wb.create_sheet("Disease-Scale Map")
    headers = [
        ("ds_map_id", 42),
        ("disease_id", 38),
        ("disease_name", 30),
        ("scale_id", 28),
        ("scale_code", 20),
        ("display_order", 14),
        ("is_required", 12),
    ]
    _write_header(ws, headers)

    row_idx = 2
    fills = [C["green_fill"], C["blue_fill"], C["purple_fill"],
             C["yellow_fill"], C["orange_fill"], C["red_fill"]]

    for di, (_, cond) in enumerate(conditions["conditions"].items()):
        name   = cond["label"]
        d_id   = disease_id(name)
        fill   = fills[di % len(fills)]
        for order, sc_code in enumerate(cond.get("scales", []), 1):
            _write_row(ws, row_idx, [
                ds_map_id(name, sc_code),
                d_id,
                name,
                scale_id(sc_code),
                sc_code,
                order,
                True,
            ], fill_hex=fill)
            row_idx += 1

    print(f"  Disease-Scale Map sheet: {row_idx - 2} rows")


def build_questions_sheet(wb, scales: dict):
    ws = wb.create_sheet("Questions")
    headers = [
        ("question_id", 28),
        ("question_code", 28),
        ("scale_id", 26),
        ("scale_code", 18),
        ("display_order", 14),
        ("question_text", 72),
        ("short_label", 28),
        ("answer_type", 16),
        ("is_required", 12),
        ("scored_in_total", 15),
        ("instructions", 40),
    ]
    _write_header(ws, headers)

    row_idx = 2
    alt = False
    for code, sc in scales.items():
        sid = scale_id(code)
        for idx, q in enumerate(sc.get("questions", [])):
            q_id = question_id(code, idx)
            fill = C["alt_fill"] if alt else None
            text = (
                q.get("question") or
                q.get("label") or
                q.get("text") or
                f"Question {idx + 1}"
            )
            label = q.get("label", "")
            # scored_in_total defaults to True unless explicitly False
            scored = q.get("scoredInTotal", True)
            _write_row(ws, row_idx, [
                q_id,
                q_id,
                sid,
                code,
                idx,
                text,
                label if label != text else "",
                q.get("type", "likert"),
                q.get("required", True),
                scored,
                q.get("instructions", ""),
            ], fill_hex=fill)
            row_idx += 1
        alt = not alt   # alternate fill per scale

    print(f"  Questions sheet: {row_idx - 2} rows")


def build_options_sheet(wb, scales: dict):
    ws = wb.create_sheet("Options")
    headers = [
        ("option_id", 32),
        ("question_id", 28),
        ("scale_code", 16),
        ("question_index", 14),
        ("option_label", 40),
        ("option_value", 14),
        ("points", 10),
        ("display_order", 14),
        ("status", 10),
    ]
    _write_header(ws, headers)

    row_idx = 2
    alt = False
    for code, sc in scales.items():
        for idx, q in enumerate(sc.get("questions", [])):
            q_id = question_id(code, idx)
            opts = q.get("options", [])
            for opt_idx, opt in enumerate(opts):
                o_id  = option_id(q_id, opt_idx)
                fill  = C["alt_fill"] if alt else None
                val   = str(opt.get("value", opt_idx))
                pts   = opt.get("points", opt.get("value", 0))
                _write_row(ws, row_idx, [
                    o_id,
                    q_id,
                    code,
                    idx,
                    opt.get("label", val),
                    val,
                    pts,
                    opt_idx,
                    True,
                ], fill_hex=fill)
                row_idx += 1
            alt = not alt

    print(f"  Options sheet: {row_idx - 2} rows")


def build_severity_bands_sheet(wb, scales: dict):
    ws = wb.create_sheet("Severity Bands")
    headers = [
        ("band_id", 32),
        ("scale_id", 26),
        ("scale_code", 18),
        ("min_score", 12),
        ("max_score", 12),
        ("level", 22),
        ("label", 38),
        ("description", 70),
    ]
    _write_header(ws, headers)

    LEVEL_FILLS = {
        "minimal":          C["green_fill"],
        "normal":           C["green_fill"],
        "mild":             C["yellow_fill"],
        "moderate":         C["orange_fill"],
        "moderately-severe": C["orange_fill"],
        "severe":           C["red_fill"],
        "borderline-high":  C["yellow_fill"],
        "high":             C["red_fill"],
    }

    row_idx = 2
    for code, sc in scales.items():
        sid = scale_id(code)
        for band in sc.get("severityBands", []):
            level = band.get("level", "")
            fill  = LEVEL_FILLS.get(level, C["alt_fill"])
            band_id = f"{code}/{level}/2026"
            _write_row(ws, row_idx, [
                band_id,
                sid,
                code,
                band.get("min", ""),
                band.get("max", ""),
                level,
                band.get("label", ""),
                band.get("description", ""),
            ], fill_hex=fill)
            row_idx += 1

    print(f"  Severity Bands sheet: {row_idx - 2} rows")


def build_summary_sheet(wb, scales: dict, conditions: dict):
    """A cover/summary sheet with stats and legend."""
    ws = wb.create_sheet("Summary", 0)   # insert at position 0

    # Title
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = "NeuroWellness PRS Data — Excel Import Reference"
    title_cell.font  = Font(name="Calibri", bold=True, size=16, color="4F46E5")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # Stats
    total_q    = sum(len(sc.get("questions", [])) for sc in scales.values())
    total_opts = sum(
        len(q.get("options", []))
        for sc in scales.values()
        for q in sc.get("questions", [])
    )
    total_maps = sum(len(c.get("scales", [])) for c in conditions["conditions"].values())

    stats = [
        ("", ""),
        ("Diseases / Conditions",    len(conditions["conditions"])),
        ("Scales (assessment tools)", len(scales)),
        ("Disease → Scale mappings",  total_maps),
        ("Questions (across all scales)", total_q),
        ("Options (answer choices)",  total_opts),
        ("Scales with severity bands",
         sum(1 for sc in scales.values() if sc.get("severityBands"))),
    ]
    for i, (label, val) in enumerate(stats, 2):
        ws.cell(row=i, column=1, value=label).font = _font(bold=True)
        ws.cell(row=i, column=2, value=val).font   = _font()

    # Sheet legend
    legend_start = len(stats) + 4
    ws.cell(row=legend_start, column=1, value="Sheet Guide").font = _font(bold=True, size=11)
    sheet_legend = [
        ("Diseases",           "prs_diseases table — one row per disease/condition"),
        ("Scales",             "prs_scales table — one row per assessment tool"),
        ("Disease-Scale Map",  "prs_disease_scale_map — links diseases to their scales"),
        ("Questions",          "prs_questions table — one row per question"),
        ("Options",            "prs_options table — one row per answer choice"),
        ("Severity Bands",     "Scoring thresholds (used by the scoring engine)"),
    ]
    for j, (sheet, desc) in enumerate(sheet_legend, legend_start + 1):
        ws.cell(row=j, column=1, value=sheet).font  = _font(bold=True, color="4F46E5")
        ws.cell(row=j, column=2, value=desc).font   = _font()

    # Column widths
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 60
    ws.sheet_view.showGridLines = False
    print(f"  Summary sheet: {len(scales)} scales, {total_q} questions, {total_opts} options")


# ── Main ───────────────────────────────────────────────────────────────────────
def main():
    if not SCALES_DIR.exists():
        print(f"ERROR: Scales directory not found:\n  {SCALES_DIR}")
        sys.exit(1)
    if not CONDITIONS_FILE.exists():
        print(f"ERROR: conditionMap.json not found:\n  {CONDITIONS_FILE}")
        sys.exit(1)

    print("NeuroWellness Excel Generator")
    print(f"  Scales dir  : {SCALES_DIR}")
    print(f"  Conditions  : {CONDITIONS_FILE}")
    print(f"  Output      : {OUTPUT_FILE}\n")

    scales     = load_all_scales()
    conditions = load_conditions()

    print(f"Loaded {len(scales)} scales, {len(conditions['conditions'])} conditions\n")
    print("Building sheets...")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default empty sheet

    build_summary_sheet(wb, scales, conditions)
    build_diseases_sheet(wb, conditions)
    build_scales_sheet(wb, scales, conditions)
    build_disease_scale_map_sheet(wb, conditions)
    build_questions_sheet(wb, scales)
    build_options_sheet(wb, scales)
    build_severity_bands_sheet(wb, scales)

    wb.save(OUTPUT_FILE)
    print(f"\nDone! Saved to:\n  {OUTPUT_FILE}")
    print("\nTo seed from JSON directly:")
    print("  python scripts/seed_scales.py \\")
    print(f'    --scales-dir "{SCALES_DIR}" \\')
    print(f'    --conditions-file "{CONDITIONS_FILE}"')


if __name__ == "__main__":
    main()
